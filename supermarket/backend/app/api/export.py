# -*- coding: utf-8 -*-
import os
import uuid
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from datetime import datetime, date
from typing import Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..database import get_db
from ..models import Product, Order, OrderDetail, StockRecord, Customer, Supplier, Employee
from ..schemas import ApiResponse
from ..utils.auth import require_admin
from ..models import User

router = APIRouter()

EXPORT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "exports")

# 样式定义
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _style_data(ws, row_idx, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _create_export_file(filename: str, wb: Workbook) -> str:
    os.makedirs(EXPORT_DIR, exist_ok=True)
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath


@router.get("/products")
async def export_products(
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """导出商品列表到Excel"""
    products = db.query(Product).order_by(Product.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "商品列表"

    headers = ["商品ID", "名称", "条形码", "分类", "单位", "进价", "售价",
               "库存量", "库存下限", "库存上限", "状态", "供应商", "创建时间"]
    _style_header(ws, headers)

    for idx, p in enumerate(products, 2):
        ws.cell(row=idx, column=1, value=p.id)
        ws.cell(row=idx, column=2, value=p.name)
        ws.cell(row=idx, column=3, value=p.barcode)
        ws.cell(row=idx, column=4, value=p.category_ref.name if p.category_ref else "")
        ws.cell(row=idx, column=5, value=p.unit)
        ws.cell(row=idx, column=6, value=float(p.purchase_price))
        ws.cell(row=idx, column=7, value=float(p.selling_price))
        ws.cell(row=idx, column=8, value=p.stock_quantity)
        ws.cell(row=idx, column=9, value=p.stock_lower_limit)
        ws.cell(row=idx, column=10, value=p.stock_upper_limit)
        ws.cell(row=idx, column=11, value=p.status.value if p.status else "")
        ws.cell(row=idx, column=12, value=p.supplier.name if p.supplier else "")
        ws.cell(row=idx, column=13, value=str(p.created_at))
        _style_data(ws, idx, len(headers))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    filename = f"products_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = _create_export_file(filename, wb)
    return ApiResponse(data={"url": f"/exports/{filename}", "filename": filename})


@router.get("/orders")
async def export_orders(
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """导出订单列表到Excel"""
    query = db.query(Order)
    if status:
        query = query.filter(Order.status == status)
    if start_date:
        query = query.filter(Order.order_time >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.filter(Order.order_time <= datetime.strptime(end_date, "%Y-%m-%d") + __import__("datetime").timedelta(days=1))
    orders = query.order_by(desc(Order.order_time)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "订单列表"

    headers = ["订单ID", "订单号", "顾客", "手机号", "金额", "状态",
               "支付方式", "收货人", "联系电话", "下单时间"]
    _style_header(ws, headers)

    for idx, o in enumerate(orders, 2):
        ws.cell(row=idx, column=1, value=o.id)
        ws.cell(row=idx, column=2, value=o.order_no)
        ws.cell(row=idx, column=3, value=o.customer.name if o.customer else "")
        ws.cell(row=idx, column=4, value=o.customer.phone if o.customer else "")
        ws.cell(row=idx, column=5, value=float(o.total_amount))
        ws.cell(row=idx, column=6, value=o.status.value if o.status else "")
        ws.cell(row=idx, column=7, value=o.payment_method)
        ws.cell(row=idx, column=8, value=o.receiver_name)
        ws.cell(row=idx, column=9, value=o.receiver_phone)
        ws.cell(row=idx, column=10, value=str(o.order_time))
        _style_data(ws, idx, len(headers))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18

    filename = f"orders_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = _create_export_file(filename, wb)
    return ApiResponse(data={"url": f"/exports/{filename}", "filename": filename})


@router.get("/stock-records")
async def export_stock_records(
    change_type: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin),
):
    """导出库存流水到Excel"""
    query = db.query(StockRecord)
    if change_type:
        query = query.filter(StockRecord.change_type == change_type)
    if start_date:
        query = query.filter(StockRecord.created_at >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.filter(StockRecord.created_at <= datetime.strptime(end_date, "%Y-%m-%d") + __import__("datetime").timedelta(days=1))
    records = query.order_by(desc(StockRecord.created_at)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "库存流水"

    headers = ["ID", "商品", "变动数量", "变动类型", "变动前", "变动后",
               "供应商", "操作人", "备注", "操作时间"]
    _style_header(ws, headers)

    for idx, r in enumerate(records, 2):
        ws.cell(row=idx, column=1, value=r.id)
        ws.cell(row=idx, column=2, value=r.product.name if r.product else "")
        ws.cell(row=idx, column=3, value=r.change_quantity)
        ws.cell(row=idx, column=4, value=r.change_type)
        ws.cell(row=idx, column=5, value=r.before_quantity)
        ws.cell(row=idx, column=6, value=r.after_quantity)
        ws.cell(row=idx, column=7, value=r.supplier_ref.name if r.supplier_ref else "")
        ws.cell(row=idx, column=8, value=r.operator)
        ws.cell(row=idx, column=9, value=r.remark)
        ws.cell(row=idx, column=10, value=str(r.created_at))
        _style_data(ws, idx, len(headers))

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    filename = f"stock_records_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = _create_export_file(filename, wb)
    return ApiResponse(data={"url": f"/exports/{filename}", "filename": filename})
